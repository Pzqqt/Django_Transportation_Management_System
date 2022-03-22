import itertools

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.writer.excel import save_virtual_workbook

_BD = Side(style='thin', color="000000")
_CENTER = Alignment(
    horizontal="center",
    vertical="center",
)
_RIGHT = Alignment(
    horizontal="right",
    vertical="center",
)
_FONT_TITLE = Font(name='黑体', size=18, bold=True)
_FONT_HEADER = Font(name='黑体', size=12, bold=True)
_FONT_VALUE = Font(name='等线', size=12, bold=False)

def add_all_border(cell_):
    """ 单元格添加所有细框线 """
    cell_.border = Border(left=_BD, top=_BD, right=_BD, bottom=_BD)

def cell_center(cell_):
    """ 单元格水平居中&垂直居中 """
    cell_.alignment = _CENTER

def cell_right(cell_):
    """ 单元格水平居右&垂直居中 """
    cell_.alignment = _RIGHT

def gen_workbook(title, thead, trs):
    wb = Workbook()
    ws = wb.active
    ws.append([title, ])
    ws.append(thead)
    for tr in trs:
        ws.append(tr)
    # 合并首行(标题)
    ws.merge_cells("A1:%s1" % ws[2][-1].column_letter)
    # 给所有单元格设置居中+所有边框格式
    for cell in itertools.chain.from_iterable(ws.rows):
        cell_center(cell)
        add_all_border(cell)
    # 给每行单元格设置字体
    ws['A1'].font = _FONT_TITLE
    for cell in ws[2]:
        cell.font = _FONT_HEADER
    for cell in itertools.chain.from_iterable(list(ws.rows)[2:]):
        cell.font = _FONT_VALUE
    # 适应列宽
    for col in ws.columns:
        max_len_set = set()
        for cell in col:
            cell_len = 0
            if not cell.value:
                continue
            for char in cell.value:
                if ord(char) <= 256:
                    cell_len += 1.3
                else:
                    cell_len += 2.6
            max_len_set.add(cell_len)
        ws.column_dimensions[col[1].column_letter].width = max(max_len_set)
    return save_virtual_workbook(wb)

def test():
    title_ = "Title"
    thead_ = ["Head 1", "Head 2", "Head 3"]
    trs_ = [["Value 1", "Value 2", "Value 3"] for _ in range(9)]
    return gen_workbook(title_, thead_, trs_)
